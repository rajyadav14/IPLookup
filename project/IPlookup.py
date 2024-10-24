#!/usr/bin/python3
import openpyxl
from ipwhois import IPWhois
import concurrent.futures
import logging


logging.basicConfig(filename='error.log', level=logging.DEBUG)

def ip_lookup(ip_address):
    """Fetch IP details using IPWhois."""
    try:
        obj = IPWhois(ip_address)
        results = obj.lookup_rdap()
        asn = results.get('asn', 'N/A')
        asn_name = results.get('asn_description', 'N/A')
        country = results.get('asn_country_code', 'N/A')
        state = results.get('network', {}).get('state', 'N/A')
        return {'ASN': asn, 'ASN Name': asn_name, 'State': state, 'Country': country}
    except Exception as e:
        logging.error(f"Error for {ip_address}: {e}")
        return {"error": str(e)}

def process_row(row):
    """Process a single row from the spreadsheet."""
    ip_address = row[0].value.strip()
    result = ip_lookup(ip_address)
    return row[0].row, result

if __name__ == "__main__":
    file_path = input("Enter the path to the spreadsheet: ")
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(process_row, row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)]
        for future in concurrent.futures.as_completed(futures):
            row_num, result = future.result()
            sheet.cell(row=row_num, column=2).value = result['ASN']
            sheet.cell(row=row_num, column=3).value = result['ASN Name']
            sheet.cell(row=row_num, column=4).value = result['State']
            sheet.cell(row=row_num, column=5).value = result['Country']
            print(f"Completed row {row_num}")

    save_path = input("Enter the path to save the updated spreadsheet: ")
    workbook.save(save_path)
    print(f"Spreadsheet saved to {save_path}.")
